"""
遺伝的アルゴリズム(GA)によるシフト表自動作成
================================================
「こいこいの人工知能研究室」YouTubeシリーズ #1〜#7 の内容を再現

エンコーディング:
  0 = 出勤
  1 = 休日（GAが割り当てる）
  2 = 希望休（固定、変更不可）

GAの流れ:
  1. Excelからデータ読み込み・クリーニング (#1)
  2. 初期個体群の生成（希望休を固定） (#2)
  3. 休日数の調整 (#3, #3.5)
  4. 評価関数（フィットネス） (#4)
  5. 関数化 (#5)
  6. 交叉・突然変異 (#6)
  7. 世代ループ (#7)
"""

import numpy as np
import pandas as pd
import copy
import random

# ============================================================
# 1. Excelからデータ読み込み (#1)
# ============================================================
def load_shift_data(filepath):
    """Excelファイルからシフト表データを読み込む"""
    df = pd.read_excel(filepath, sheet_name='シフト表', header=None)

    # シフト表本体: 行4〜13(0始まり)=社員A〜J, 列1〜31=1日〜31日
    shift_df = df.iloc[4:14, 1:32].copy()
    shift_df = shift_df.fillna(0)
    shift_df = shift_df.replace('◎', 2)
    shift_df = shift_df.astype(int)

    # numpy配列に変換 (10社員 x 31日)
    shift_array = shift_df.values

    # 休日数テーブル: 列32(0始まり)
    holiday_df = df.iloc[4:14, 32:33].copy()
    holiday_df = holiday_df.reset_index(drop=True)
    holiday_counts = holiday_df.iloc[:, 0].astype(int).values

    # 社員名
    employee_names = df.iloc[4:14, 0].values.tolist()

    # 必要出勤人数: 行15, 列1〜31
    required_workers = df.iloc[15, 1:32].astype(int).values

    # 列名を日付に
    days = list(range(1, 32))

    return shift_array, holiday_counts, employee_names, required_workers, days


# ============================================================
# 2. 初期個体群の生成 (#2)
# ============================================================
def create_individual(shift_template, holiday_counts):
    """
    1個体（シフト表）を生成する。
    希望休(2)は固定し、残りのセルに休日(1)をランダムに配置。
    各社員の休日数が契約休日数になるように調整。
    """
    individual = shift_template.copy()
    num_employees, num_days = individual.shape

    for emp_idx in range(num_employees):
        # この社員の希望休の数
        preferred_count = np.sum(individual[emp_idx] == 2)
        # 必要な追加休日数
        additional_holidays = holiday_counts[emp_idx] - preferred_count

        if additional_holidays <= 0:
            continue

        # 出勤(0)のセルのインデックス
        work_indices = np.where(individual[emp_idx] == 0)[0]

        if len(work_indices) < additional_holidays:
            additional_holidays = len(work_indices)

        # ランダムに休日を割り当て
        chosen = np.random.choice(work_indices, size=additional_holidays, replace=False)
        individual[emp_idx, chosen] = 1

    return individual


def create_population(shift_template, holiday_counts, pop_size):
    """初期個体群を生成"""
    population = []
    for _ in range(pop_size):
        ind = create_individual(shift_template, holiday_counts)
        population.append(ind)
    return population


# ============================================================
# 3. 休日数の調整 (#3, #3.5)
# ============================================================
def adjust_holidays(individual, holiday_counts):
    """
    各社員の休日数（1と2の合計）が契約休日数と一致するように調整。
    多い場合：ランダムに休日(1)を出勤(0)に戻す
    少ない場合：ランダムに出勤(0)を休日(1)にする
    """
    adjusted = individual.copy()
    num_employees = adjusted.shape[0]

    for emp_idx in range(num_employees):
        current_holidays = np.sum((adjusted[emp_idx] == 1) | (adjusted[emp_idx] == 2))
        target = holiday_counts[emp_idx]

        if current_holidays > target:
            # 休日が多すぎる → 休日(1)を出勤(0)に戻す（希望休2は変更しない）
            excess = current_holidays - target
            holiday_indices = np.where(adjusted[emp_idx] == 1)[0]
            if len(holiday_indices) >= excess:
                remove = np.random.choice(holiday_indices, size=excess, replace=False)
                adjusted[emp_idx, remove] = 0

        elif current_holidays < target:
            # 休日が足りない → 出勤(0)を休日(1)に
            deficit = target - current_holidays
            work_indices = np.where(adjusted[emp_idx] == 0)[0]
            if len(work_indices) >= deficit:
                add = np.random.choice(work_indices, size=deficit, replace=False)
                adjusted[emp_idx, add] = 1

    return adjusted


# ============================================================
# 4. 評価関数 (#4)
# ============================================================
def evaluate(individual, holiday_counts, required_workers):
    """
    個体の適合度（フィットネス）を計算。
    ペナルティ方式: 値が小さいほど良い（0が最良）。

    ペナルティ項目:
    1. 各日の出勤人数と必要人数の差（重みつき）
    2. 各社員の休日数と契約休日数の差
    3. 連続勤務日数が多い場合のペナルティ
    """
    penalty = 0
    num_employees, num_days = individual.shape

    # --- ペナルティ1: 日ごとの出勤人数の過不足 ---
    for day in range(num_days):
        # 出勤者数（0のセル数）
        workers = np.sum(individual[:, day] == 0)
        diff = abs(workers - required_workers[day])
        penalty += diff * 10  # 重み: 10

    # --- ペナルティ2: 社員ごとの休日数の過不足 ---
    for emp_idx in range(num_employees):
        total_holidays = np.sum((individual[emp_idx] == 1) | (individual[emp_idx] == 2))
        diff = abs(total_holidays - holiday_counts[emp_idx])
        penalty += diff * 20  # 重み: 20（休日数は厳守させたい）

    # --- ペナルティ3: 連続勤務ペナルティ ---
    for emp_idx in range(num_employees):
        consecutive = 0
        for day in range(num_days):
            if individual[emp_idx, day] == 0:
                consecutive += 1
                if consecutive > 5:  # 6日以上連続勤務はペナルティ
                    penalty += (consecutive - 5) * 5
            else:
                consecutive = 0

    return penalty


# ============================================================
# 5. 選択（トーナメント選択）
# ============================================================
def tournament_selection(population, fitness_list, tournament_size=3):
    """トーナメント選択で1個体を選ぶ"""
    indices = random.sample(range(len(population)), tournament_size)
    best_idx = min(indices, key=lambda i: fitness_list[i])
    return population[best_idx].copy()


# ============================================================
# 6. 交叉・突然変異 (#6)
# ============================================================
def crossover(parent1, parent2, shift_template):
    """
    一様交叉: 各社員について、ランダムに親1か親2の遺伝子を選択。
    希望休(2)は固定のまま維持。
    """
    child = parent1.copy()
    num_employees, num_days = child.shape

    for emp_idx in range(num_employees):
        # 各日について50%の確率で親2の遺伝子を採用
        mask = np.random.random(num_days) < 0.5
        # ただし希望休のセルは変更しない
        preferred_mask = shift_template[emp_idx] == 2
        mask = mask & ~preferred_mask
        child[emp_idx, mask] = parent2[emp_idx, mask]

    return child


def mutate(individual, shift_template, mutation_rate=0.02):
    """
    突然変異: 各セルについて一定確率で出勤⇔休日を入れ替え。
    希望休(2)のセルは変更しない。
    """
    mutated = individual.copy()
    num_employees, num_days = mutated.shape

    for emp_idx in range(num_employees):
        for day in range(num_days):
            if shift_template[emp_idx, day] == 2:
                continue  # 希望休は変更しない
            if random.random() < mutation_rate:
                # 出勤(0)と休日(1)を入れ替え
                if mutated[emp_idx, day] == 0:
                    mutated[emp_idx, day] = 1
                else:
                    mutated[emp_idx, day] = 0

    return mutated


# ============================================================
# 7. 世代ループ（メインGA） (#7)
# ============================================================
def run_ga(shift_template, holiday_counts, required_workers,
           pop_size=100, generations=500, mutation_rate=0.02,
           elite_size=5, tournament_size=3):
    """
    遺伝的アルゴリズムのメインループ

    Parameters:
    -----------
    shift_template : np.array
        初期シフトテンプレート（希望休=2が入っている）
    holiday_counts : np.array
        各社員の契約休日数
    required_workers : np.array
        各日の必要出勤人数
    pop_size : int
        個体群サイズ
    generations : int
        世代数
    mutation_rate : float
        突然変異率
    elite_size : int
        エリート数（そのまま次世代に残す個体数）
    tournament_size : int
        トーナメントサイズ

    Returns:
    --------
    best_individual : np.array
        最良のシフト表
    best_fitness : float
        最良の適合度
    history : list
        各世代の最良適合度の履歴
    """
    print(f"=== GA開始 ===")
    print(f"個体群サイズ: {pop_size}, 世代数: {generations}")
    print(f"突然変異率: {mutation_rate}, エリート数: {elite_size}")
    print()

    # 初期個体群生成
    population = create_population(shift_template, holiday_counts, pop_size)

    # 適合度計算
    fitness_list = [evaluate(ind, holiday_counts, required_workers) for ind in population]

    best_fitness = min(fitness_list)
    best_idx = fitness_list.index(best_fitness)
    best_individual = population[best_idx].copy()
    history = [best_fitness]

    print(f"初期世代: 最良適合度 = {best_fitness}")

    for gen in range(1, generations + 1):
        # --- エリート選択 ---
        sorted_indices = sorted(range(len(fitness_list)), key=lambda i: fitness_list[i])
        elites = [population[i].copy() for i in sorted_indices[:elite_size]]

        # --- 次世代生成 ---
        new_population = list(elites)  # エリートをそのまま残す

        while len(new_population) < pop_size:
            # 親選択
            parent1 = tournament_selection(population, fitness_list, tournament_size)
            parent2 = tournament_selection(population, fitness_list, tournament_size)

            # 交叉
            child = crossover(parent1, parent2, shift_template)

            # 突然変異
            child = mutate(child, shift_template, mutation_rate)

            # 休日数調整
            child = adjust_holidays(child, holiday_counts)

            new_population.append(child)

        # 次世代に更新
        population = new_population
        fitness_list = [evaluate(ind, holiday_counts, required_workers) for ind in population]

        # 最良個体の更新
        gen_best_fitness = min(fitness_list)
        if gen_best_fitness < best_fitness:
            best_fitness = gen_best_fitness
            best_idx = fitness_list.index(best_fitness)
            best_individual = population[best_idx].copy()

        history.append(gen_best_fitness)

        # 進捗表示
        if gen % 50 == 0 or gen == 1:
            print(f"世代 {gen:4d}: 最良適合度 = {gen_best_fitness:6.1f} "
                  f"(全体最良 = {best_fitness:6.1f})")

        # 最適解に到達
        if best_fitness == 0:
            print(f"\n*** 最適解発見! 世代 {gen} ***")
            break

    print(f"\n=== GA完了 ===")
    print(f"最終最良適合度: {best_fitness}")

    return best_individual, best_fitness, history


# ============================================================
# 結果をExcelに出力
# ============================================================
def save_result_to_excel(result, employee_names, days, holiday_counts,
                         required_workers, filepath):
    """GAの結果をExcelファイルに保存"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "GA結果シフト表"

    # スタイル
    header_font = Font(bold=True, size=11, name='Arial')
    center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    blue_fill = PatternFill('solid', fgColor='DAEEF3')
    green_fill = PatternFill('solid', fgColor='E2EFDA')
    yellow_fill = PatternFill('solid', fgColor='FFFF00')
    red_font = Font(color='FF0000', bold=True, name='Arial')
    work_fill = PatternFill('solid', fgColor='FFFFFF')
    holiday_fill = PatternFill('solid', fgColor='D9E2F3')
    preferred_fill = PatternFill('solid', fgColor='FCE4EC')

    # タイトル
    ws.cell(row=1, column=1, value='GA最適化シフト表')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, name='Arial')

    # ヘッダー
    ws.cell(row=3, column=1, value='社員名').font = header_font
    ws.cell(row=3, column=1).fill = blue_fill
    ws.cell(row=3, column=1).alignment = center
    ws.cell(row=3, column=1).border = border

    for i, day in enumerate(days):
        cell = ws.cell(row=3, column=i + 2, value=day)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    # 休日数ヘッダー
    cell = ws.cell(row=3, column=33, value='休日数')
    cell.font = header_font
    cell.fill = yellow_fill
    cell.alignment = center
    cell.border = border

    cell = ws.cell(row=3, column=34, value='契約')
    cell.font = header_font
    cell.fill = yellow_fill
    cell.alignment = center
    cell.border = border

    # データ
    label_map = {0: '出', 1: '休', 2: '◎'}

    for emp_idx, name in enumerate(employee_names):
        row = 4 + emp_idx
        ws.cell(row=row, column=1, value=name).font = Font(name='Arial', size=11)
        ws.cell(row=row, column=1).fill = green_fill
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).border = border

        actual_holidays = 0
        for day_idx in range(len(days)):
            val = result[emp_idx, day_idx]
            cell = ws.cell(row=row, column=day_idx + 2)
            cell.value = label_map[val]
            cell.alignment = center
            cell.border = border

            if val == 0:
                cell.fill = work_fill
            elif val == 1:
                cell.fill = holiday_fill
                actual_holidays += 1
            elif val == 2:
                cell.fill = preferred_fill
                cell.font = red_font
                actual_holidays += 1

        # 実休日数
        cell = ws.cell(row=row, column=33, value=actual_holidays)
        cell.alignment = center
        cell.border = border
        # 契約休日数
        cell = ws.cell(row=row, column=34, value=holiday_counts[emp_idx])
        cell.alignment = center
        cell.border = border
        cell.fill = yellow_fill

    # 出勤人数集計行
    summary_row = 4 + len(employee_names) + 1
    ws.cell(row=summary_row, column=1, value='出勤人数').font = header_font
    ws.cell(row=summary_row, column=1).fill = PatternFill('solid', fgColor='FFC000')
    ws.cell(row=summary_row, column=1).alignment = center
    ws.cell(row=summary_row, column=1).border = border

    for day_idx in range(len(days)):
        workers = np.sum(result[:, day_idx] == 0)
        cell = ws.cell(row=summary_row, column=day_idx + 2, value=int(workers))
        cell.alignment = center
        cell.border = border
        if workers < required_workers[day_idx]:
            cell.fill = PatternFill('solid', fgColor='FF9999')  # 不足は赤
        elif workers > required_workers[day_idx]:
            cell.fill = PatternFill('solid', fgColor='FFFF99')  # 超過は黄
        else:
            cell.fill = PatternFill('solid', fgColor='99FF99')  # 適正は緑

    # 必要人数行
    req_row = summary_row + 1
    ws.cell(row=req_row, column=1, value='必要人数').font = header_font
    ws.cell(row=req_row, column=1).fill = PatternFill('solid', fgColor='FFC000')
    ws.cell(row=req_row, column=1).alignment = center
    ws.cell(row=req_row, column=1).border = border
    for day_idx in range(len(days)):
        cell = ws.cell(row=req_row, column=day_idx + 2, value=int(required_workers[day_idx]))
        cell.alignment = center
        cell.border = border
        cell.fill = PatternFill('solid', fgColor='FFC000')

    # 列幅
    ws.column_dimensions['A'].width = 12
    for col_idx in range(2, 35):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    ws.column_dimensions[openpyxl.utils.get_column_letter(33)].width = 8
    ws.column_dimensions[openpyxl.utils.get_column_letter(34)].width = 6

    from openpyxl.utils import get_column_letter
    wb.save(filepath)
    print(f"結果を保存: {filepath}")


# ============================================================
# メイン実行
# ============================================================
if __name__ == '__main__':
    import openpyxl

    # 1. データ読み込み
    print("=" * 60)
    print("遺伝的アルゴリズムによるシフト表自動作成")
    print("=" * 60)
    print()

    input_file = '/sessions/kind-gifted-clarke/shift_input.xlsx'
    output_file = '/sessions/kind-gifted-clarke/shift_result.xlsx'

    shift_template, holiday_counts, employee_names, required_workers, days = \
        load_shift_data(input_file)

    print(f"社員数: {len(employee_names)}")
    print(f"日数: {len(days)}")
    print(f"社員名: {employee_names}")
    print(f"契約休日数: {holiday_counts}")
    print(f"必要出勤人数: {required_workers[0]}人/日")
    print()

    # 希望休の状況表示
    print("【希望休の状況】")
    for i, name in enumerate(employee_names):
        preferred_days = np.where(shift_template[i] == 2)[0] + 1
        print(f"  {name}: {list(preferred_days)}日")
    print()

    # 2. GA実行
    best_shift, best_fitness, history = run_ga(
        shift_template=shift_template,
        holiday_counts=holiday_counts,
        required_workers=required_workers,
        pop_size=100,
        generations=500,
        mutation_rate=0.02,
        elite_size=5,
        tournament_size=3
    )

    # 3. 結果の確認
    print()
    print("=" * 60)
    print("【結果の確認】")
    print("=" * 60)

    # 各日の出勤人数
    print("\n日ごとの出勤人数:")
    for day_idx in range(len(days)):
        workers = np.sum(best_shift[:, day_idx] == 0)
        status = "✓" if workers == required_workers[day_idx] else "✗"
        print(f"  {days[day_idx]:2d}日: {workers}人 (必要: {required_workers[day_idx]}人) {status}")

    # 各社員の休日数
    print("\n社員ごとの休日数:")
    for emp_idx, name in enumerate(employee_names):
        total_holidays = np.sum((best_shift[emp_idx] == 1) | (best_shift[emp_idx] == 2))
        status = "✓" if total_holidays == holiday_counts[emp_idx] else "✗"
        print(f"  {name}: {total_holidays}日 (契約: {holiday_counts[emp_idx]}日) {status}")

    # 4. Excelに出力
    save_result_to_excel(best_shift, employee_names, days, holiday_counts,
                         required_workers, output_file)

    # 5. 適合度推移のサマリー
    print(f"\n適合度推移: {history[0]} → {history[-1]}")
    print(f"改善率: {(1 - history[-1]/max(history[0],1))*100:.1f}%")
