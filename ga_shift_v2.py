"""
遺伝的アルゴリズム(GA)によるシフト表自動作成
================================================
「こいこいの人工知能研究室」YouTubeシリーズ #1〜#7 の字幕に基づく忠実な再現

字幕分析により判明した実装の特徴:
- 選択方式: 切り捨て選択（上位N個体を残す）※トーナメント選択ではない
- 交叉: 一様交叉（同じ遺伝子はそのまま、異なる場合50%で振り分け）
- 突然変異: 5%の確率で発生し、発生時は10%の遺伝子をランダム変異
- 個体構造: (スコア, pandasデータフレーム) のタプルリスト
- 評価: 文字列分割によるパターンマッチング（split）
- 世代間最強個体の保存（エリート保存）
- 休日生成: whileループで重複チェック（np.random.choiceではない）
- 1次元化: .values.flatten() で交叉・突然変異を行う

エンコーディング:
  0 = 出勤
  1 = 休日（GAが割り当て）
  2 = 希望休（固定、変更不可）
"""

import numpy as np
import pandas as pd
import random
import itertools

# pandas FutureWarning抑制
pd.set_option('future.no_silent_downcasting', True)


# ============================================================
# 第1回: Excelからデータ読み込み (read_xl)
# ============================================================
def read_xl(filepath='shift_input.xlsx', sheet_name='シフト表'):
    """
    Excelファイルからシフト表データを読み込む。
    動画#1の実装に忠実:
    - pd.read_excel() で読み込み
    - iloc[4:14, 1:32] でシフト本体を抜き取り
    - fillna(0) でNaN→0
    - replace('◎', 2) で希望休→2
    - 休日数テーブル: iloc[4:14, 32:33]
    - 列名をリスト内包表記で1〜31に変更
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)

    # シフト表本体を抜き取る (10人 x 31日)
    kiso = df.iloc[4:14, 1:32].copy()
    kiso = kiso.fillna(0).infer_objects(copy=False)
    kiso = kiso.replace('◎', 2).infer_objects(copy=False)

    # 列名を1〜31に変更（内包表記）
    kiso.columns = [i + 1 for i in range(len(kiso.columns))]
    kiso = kiso.reset_index(drop=True)

    # 休日数テーブル
    holiday = df.iloc[4:14, 32:33].copy()
    holiday = holiday.reset_index(drop=True)
    holiday.columns = ['休日数']

    # 必要出勤人数（行15, 列1〜31）
    required = df.iloc[15, 1:32].copy()
    required = required.fillna(7).infer_objects(copy=False).astype(int).values

    # 社員名
    employee_names = df.iloc[4:14, 0].values.tolist()

    return kiso, holiday, required, employee_names


# ============================================================
# 第2回: 第1世代の個体生成 (first_gene)
# ============================================================
def first_gene(kiso, holiday):
    """
    第1世代の1個体を生成する。
    動画#2の実装に忠実:
    - kiso をコピー
    - 各社員ごとにwhileループで重複なしのランダム休日を生成
    - np.random.randint(1, days+1) で1つずつ乱数生成
    - h リストに重複チェックして追加
    - loc[行, 列] でセルにアクセス
    - 0のセルだけ1に変更、2（希望休）はスキップ
    """
    kiso_copy = kiso.copy()
    days = len(kiso_copy.columns)  # 31

    for k in range(len(kiso_copy)):
        # この社員の休日数
        target_holidays = int(holiday.loc[k, '休日数'])

        # 重複なしのランダム休日リストを生成
        h = []
        while len(h) < target_holidays:
            n = np.random.randint(1, days + 1)
            if n not in h:
                h.append(n)

        # 休日を埋め込む
        for n in h:
            if kiso_copy.loc[k, n] == 0:
                kiso_copy.loc[k, n] = 1
            # 2（希望休）の場合は何もしない

    return kiso_copy


# ============================================================
# 第3回+3.5回: 休日数の調整 (holiday_fix)
# ============================================================
def holiday_fix(kiso_copy, holiday):
    """
    各社員の休日数を契約休日数に合わせて調整する。
    動画#3, #3.5の実装に忠実:
    - np.count_nonzero で休日数（1と2）をカウント
    - 差分 s = 実際 - 契約
    - s > 0: 休日が多い → 1を0に変更
    - s < 0: 休日が少ない → 0を1に変更
    - #3.5で統合: abs(s)でループ、c/c1変数で方向制御
    - whileループで該当セルをランダムに見つけて変更
    """
    days = len(kiso_copy.columns)

    for k in range(len(kiso_copy)):
        # 休日数カウント（0以外=1と2の合計）
        actual = np.count_nonzero(kiso_copy.iloc[k].values)
        target = int(holiday.loc[k, '休日数'])
        s = actual - target

        if s == 0:
            continue

        # #3.5の統合版: s > 0なら減らす、s < 0なら増やす
        s_abs = abs(s)
        if s > 0:
            c = 1   # 探す値（1=休日を見つけて）
            c1 = 0  # 変更先（0=出勤に戻す）
        else:
            c = 0   # 探す値（0=出勤を見つけて）
            c1 = 1  # 変更先（1=休日にする）

        # 休日を増減したい数に達するまでループ
        buf = 0
        while buf < s_abs:
            n = np.random.randint(1, days + 1)
            if kiso_copy.loc[k, n] == c:
                kiso_copy.loc[k, n] = c1
                buf += 1

    return kiso_copy


# ============================================================
# 第4回: 評価関数 (evaluation_function)
# ============================================================
def evaluation_function(kiso_copy, required_workers=None):
    """
    シフト表の評価関数。ペナルティ方式（0に近いほど良い）。
    動画#4の実装に忠実:
    - まず希望休(2)を1に変換（評価時は区別不要）
    - 文字列に変換してsplit()によるパターンマッチング

    評価項目:
    1. 5連勤以上: split('1')で区切り、0の連続長が5以上なら -(長さ-4)^2
    2. 3連休以上: split('0')で区切り、1の連続長が3以上なら加点（今回は省略してペナルティ0）
    3. 飛び石連休: '1 0 1'で区切り、(分割数-1) * -10
    4. 出勤人数: 各日の出勤数と必要人数の差 * -4
    """
    # 希望休(2)を1に変換（評価時は1も2も「休日」として同一視）
    eval_copy = kiso_copy.copy()
    eval_copy = eval_copy.replace(2, 1)

    score = 0

    # --- 横方向の評価（各社員ごと） ---
    for k in range(len(eval_copy)):
        # 1次元化して文字列に変換
        row_str = ' '.join([str(int(x)) for x in eval_copy.iloc[k].values.flatten()])

        # (1) 5連勤以上の評価
        # '1'(休日)で分割すると、出勤の連続部分が取れる
        segments = row_str.split('1')
        for seg in segments:
            # 各セグメントの0の数を数える（スペース区切りなので）
            zeros = seg.strip().split()
            consecutive_work = len([x for x in zeros if x == '0'])
            if consecutive_work >= 5:
                # -(連勤数 - 4)^2 のペナルティ
                penalty = -((consecutive_work - 4) ** 2)
                score += penalty

        # (2) 3連休以上の評価
        segments_holiday = row_str.split('0')
        for seg in segments_holiday:
            ones = seg.strip().split()
            consecutive_rest = len([x for x in ones if x == '1'])
            if consecutive_rest >= 3:
                # 3連休以上は（今回は）特にペナルティなし
                # 動画では説明を割愛していたため、0点とする
                pass

        # (3) 飛び石連休の評価
        # '1 0 1' パターンで分割
        tobishi_parts = row_str.split('1 0 1')
        tobishi_count = len(tobishi_parts) - 1
        score += tobishi_count * (-10)

    # --- 縦方向の評価（日ごとの出勤人数） ---
    if required_workers is None:
        # デフォルト: 全体の7割（10人なら7人）
        target = int(len(eval_copy) * 0.7)
        required_workers = np.full(len(eval_copy.columns), target)

    for day_idx, day in enumerate(eval_copy.columns):
        # 各日の休日合計（1の数）
        holiday_sum = int(eval_copy[day].sum())
        # 出勤人数 = 全体 - 休日数
        workers = len(eval_copy) - holiday_sum
        # 必要人数との差
        diff = abs(workers - required_workers[day_idx])
        score += diff * (-4)

    return score


# ============================================================
# 第6回: 一様交叉 (crossover)
# ============================================================
def crossover(parent1_df, parent2_df, ep=0.5):
    """
    一様交叉。
    動画#6の実装に忠実:
    - 両親を1次元化 (.values.flatten())
    - 同じ遺伝子はそのまま継承
    - 異なる遺伝子は ep(50%)の確率で振り分け
    - 2つの子を生成
    - zip(p1, p2) でペアで走査
    """
    p1 = parent1_df.values.flatten()
    p2 = parent2_df.values.flatten()

    ch1 = []
    ch2 = []

    for v1, v2 in zip(p1, p2):
        if v1 == v2:
            # 同じ → そのまま継承
            ch1.append(v1)
            ch2.append(v2)
        else:
            # 異なる → 50%で振り分け
            if random.random() < ep:
                ch1.append(v1)
                ch2.append(v2)
            else:
                ch1.append(v2)
                ch2.append(v1)

    return np.array(ch1), np.array(ch2)


# ============================================================
# 第6回: 突然変異 (mutation)
# ============================================================
def mutation(child_1d, sd=0.05):
    """
    突然変異。
    動画#6+#7修正の実装に忠実:
    - sd(5%)の確率で突然変異が発生
    - 発生時: np.random.permutation で全インデックスをシャッフル
    - 先頭10%のインデックスの遺伝子を変異
    - 0→1, 1→0 に変更（2=希望休は変更しない）
    - #7の修正: elseではなくif文で0と1を明示的にチェック
    """
    ch = child_1d.copy()

    if random.random() < sd:
        # 突然変異発生
        indices = np.random.permutation(len(ch))
        mutation_count = int(len(ch) * 0.1)  # 10%の遺伝子を変異
        target_indices = indices[:mutation_count]

        for idx in target_indices:
            # #7で修正されたロジック: if文で明示チェック（elseだと2も変わる）
            if ch[idx] == 1:
                ch[idx] = 0
            elif ch[idx] == 0:
                ch[idx] = 1
            # 2（希望休）はスキップ

    return ch


# ============================================================
# 1次元→2次元変換ユーティリティ
# ============================================================
def to_dataframe(child_1d, kiso_template):
    """1次元配列をpandasデータフレームに戻す"""
    rows = len(kiso_template)
    cols = len(kiso_template.columns)
    data = child_1d.reshape(rows, cols)
    df = pd.DataFrame(data, columns=kiso_template.columns)
    return df


# ============================================================
# 第7回: メインGA世代ループ (run_ga)
# ============================================================
def run_ga(filepath='shift_input.xlsx',
           elite_length=20,
           generation_count=50,
           ep=0.5,
           sd=0.05):
    """
    遺伝的アルゴリズムのメインループ。
    動画#7の実装に忠実:

    1. 第1世代を100個体生成
    2. 全個体を評価・点数順にソート
    3. 上位elite_length個体を選択
    4. 全世代最強個体(top)を保存・比較
    5. 上位個体の全組み合わせ(C(n,2))で交叉→子生成
    6. 子に突然変異→休日数調整→評価
    7. 次世代に進む
    8. generation_count世代繰り返す

    個体構造: parent = [(score, dataframe), ...]
    選択方式: 切り捨て選択（上位N個体）※トーナメントではない
    """
    print("=" * 60)
    print("遺伝的アルゴリズムによるシフト表自動作成")
    print("=" * 60)

    # データ読み込み
    kiso, holiday, required_workers, employee_names = read_xl(filepath)

    print(f"社員数: {len(kiso)}")
    print(f"日数: {len(kiso.columns)}")
    print(f"社員名: {employee_names}")
    print(f"契約休日数: {list(holiday['休日数'].values)}")
    print(f"エリート数: {elite_length}, 世代数: {generation_count}")
    print(f"交叉確率: {ep}, 突然変異率: {sd}")
    print()

    # 希望休の状況表示
    print("【希望休の状況】")
    for k in range(len(kiso)):
        preferred = [col for col in kiso.columns if kiso.loc[k, col] == 2]
        print(f"  {employee_names[k]}: {preferred}日")
    print()

    # === 1. 第1世代を100個体生成 ===
    initial_pop = 100
    parent = []
    for _ in range(initial_pop):
        ind = first_gene(kiso, holiday)
        ind = holiday_fix(ind, holiday)
        sc = evaluation_function(ind, required_workers)
        parent.append((sc, ind))

    # 全世代最強を追跡する変数
    top = None
    history = []

    # === 世代ループ ===
    for gen in range(generation_count):

        # === 2. 点数順にソート（0に近い＝良いスコアが上位） ===
        parent = sorted(parent, key=lambda x: -x[0])

        # === 3. 上位elite_length個体を選択 ===
        parent = parent[:elite_length]

        # === 4. 全世代最強個体(top)の保存・比較 ===
        if gen == 0:
            top = parent[0]
        elif parent[0][0] > top[0]:
            top = parent[0]
        else:
            parent.append(top)

        # 進捗表示
        print(f"第{gen + 1:3d}世代: 最高点 = {parent[0][0]:6.0f}  "
              f"(全世代最強 = {top[0]:6.0f})")
        history.append(top[0])

        if gen == 0:
            print(f"  初期最良シフト:")
            print(f"  {parent[0][1].values}")
            print()

        # === 5. 上位個体の全組み合わせで交叉 ===
        children = []
        for k1 in range(len(parent)):
            for k2 in range(k1 + 1, len(parent)):
                p1_df = parent[k1][1]
                p2_df = parent[k2][1]

                ch1_1d, ch2_1d = crossover(p1_df, p2_df, ep)

                ch1_1d = mutation(ch1_1d, sd)
                ch2_1d = mutation(ch2_1d, sd)

                ch1_df = to_dataframe(ch1_1d, kiso)
                ch2_df = to_dataframe(ch2_1d, kiso)

                ch1_df = holiday_fix(ch1_df, holiday)
                ch2_df = holiday_fix(ch2_df, holiday)

                sc1 = evaluation_function(ch1_df, required_workers)
                sc2 = evaluation_function(ch2_df, required_workers)

                children.append((sc1, ch1_df))
                children.append((sc2, ch2_df))

        # === 6. 次世代 = エリート + 子 ===
        parent = list(parent) + children

    # === 最終結果 ===
    parent = sorted(parent, key=lambda x: -x[0])
    best_score = parent[0][0]
    best_shift = parent[0][1]

    if top[0] > best_score:
        best_score = top[0]
        best_shift = top[1]

    print()
    print("=" * 60)
    print(f"=== GA完了: 最終最高点 = {best_score} ===")
    print("=" * 60)

    return best_shift, best_score, history, kiso, holiday, required_workers, employee_names


# ============================================================
# 結果をExcelに出力
# ============================================================
def save_result_to_excel(result_df, employee_names, holiday, required_workers, filepath):
    """GAの結果をExcelファイルに保存"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "GA結果シフト表"

    # スタイル
    header_font = Font(bold=True, size=11, name='Arial')
    center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    blue_fill = PatternFill('solid', fgColor='DAEEF3')
    green_fill = PatternFill('solid', fgColor='E2EFDA')
    yellow_fill = PatternFill('solid', fgColor='FFFF00')
    red_font = Font(color='FF0000', bold=True, name='Arial')
    work_fill = PatternFill('solid', fgColor='FFFFFF')
    holiday_fill = PatternFill('solid', fgColor='D9E2F3')
    preferred_fill = PatternFill('solid', fgColor='FCE4EC')

    # タイトル
    ws.cell(row=1, column=1, value='GA最適化シフト表（忠実再現版）')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, name='Arial')

    days = list(result_df.columns)

    # ヘッダー
    ws.cell(row=3, column=1, value='社員名').font = header_font
    ws.cell(row=3, column=1).fill = blue_fill
    ws.cell(row=3, column=1).alignment = center
    ws.cell(row=3, column=1).border = border

    for i, day in enumerate(days):
        cell = ws.cell(row=3, column=i + 2, value=day)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    # 休日数ヘッダー
    col_actual = len(days) + 2
    col_contract = len(days) + 3
    for col, label in [(col_actual, '実休日'), (col_contract, '契約')]:
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = center
        cell.border = border

    # データ
    label_map = {0: '出', 1: '休', 2: '◎'}

    for k, name in enumerate(employee_names):
        row = 4 + k
        ws.cell(row=row, column=1, value=name).font = Font(name='Arial', size=11)
        ws.cell(row=row, column=1).fill = green_fill
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).border = border

        actual_holidays = 0
        for day_idx, day in enumerate(days):
            val = int(result_df.loc[k, day])
            cell = ws.cell(row=row, column=day_idx + 2)
            cell.value = label_map.get(val, str(val))
            cell.alignment = center
            cell.border = border

            if val == 0:
                cell.fill = work_fill
            elif val == 1:
                cell.fill = holiday_fill
                actual_holidays += 1
            elif val == 2:
                cell.fill = preferred_fill
                cell.font = red_font
                actual_holidays += 1

        # 実休日数
        ws.cell(row=row, column=col_actual, value=actual_holidays).alignment = center
        ws.cell(row=row, column=col_actual).border = border
        # 契約休日数
        contract = int(holiday.loc[k, '休日数'])
        cell = ws.cell(row=row, column=col_contract, value=contract)
        cell.alignment = center
        cell.border = border
        cell.fill = yellow_fill

    # 出勤人数集計行
    summary_row = 4 + len(employee_names) + 1
    ws.cell(row=summary_row, column=1, value='出勤人数').font = header_font
    ws.cell(row=summary_row, column=1).fill = PatternFill('solid', fgColor='FFC000')
    ws.cell(row=summary_row, column=1).alignment = center
    ws.cell(row=summary_row, column=1).border = border

    for day_idx, day in enumerate(days):
        eval_col = result_df[day].replace(2, 1)
        holiday_count = int(eval_col.sum())
        workers = len(result_df) - holiday_count
        cell = ws.cell(row=summary_row, column=day_idx + 2, value=workers)
        cell.alignment = center
        cell.border = border
        target = required_workers[day_idx]
        if workers < target:
            cell.fill = PatternFill('solid', fgColor='FF9999')
        elif workers > target:
            cell.fill = PatternFill('solid', fgColor='FFFF99')
        else:
            cell.fill = PatternFill('solid', fgColor='99FF99')

    # 必要人数行
    req_row = summary_row + 1
    ws.cell(row=req_row, column=1, value='必要人数').font = header_font
    ws.cell(row=req_row, column=1).fill = PatternFill('solid', fgColor='FFC000')
    ws.cell(row=req_row, column=1).alignment = center
    ws.cell(row=req_row, column=1).border = border
    for day_idx in range(len(days)):
        cell = ws.cell(row=req_row, column=day_idx + 2, value=int(required_workers[day_idx]))
        cell.alignment = center
        cell.border = border
        cell.fill = PatternFill('solid', fgColor='FFC000')

    # 列幅
    ws.column_dimensions['A'].width = 12
    for col_idx in range(2, col_contract + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    ws.column_dimensions[get_column_letter(col_actual)].width = 8
    ws.column_dimensions[get_column_letter(col_contract)].width = 6

    wb.save(filepath)
    print(f"結果を保存: {filepath}")


# ============================================================
# メイン実行
# ============================================================
if __name__ == '__main__':
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, 'shift_input.xlsx')
    output_file = os.path.join(script_dir, 'shift_result_v2.xlsx')

    best_shift, best_score, history, kiso, holiday, required_workers, employee_names = \
        run_ga(
            filepath=input_file,
            elite_length=20,
            generation_count=50,
            ep=0.5,
            sd=0.05
        )

    # 結果の確認
    print()
    print("【結果の確認】")

    # 各日の出勤人数
    print("\n日ごとの出勤人数:")
    eval_shift = best_shift.replace(2, 1)
    for day in best_shift.columns:
        holiday_count = int(eval_shift[day].sum())
        workers = len(best_shift) - holiday_count
        target = required_workers[day - 1]
        status = "✓" if workers == target else "✗"
        print(f"  {day:2d}日: {workers}人 (必要: {target}人) {status}")

    # 各社員の休日数
    print("\n社員ごとの休日数:")
    for k, name in enumerate(employee_names):
        total = int(np.count_nonzero(best_shift.iloc[k].values))
        contract = int(holiday.loc[k, '休日数'])
        status = "✓" if total == contract else "✗"
        print(f"  {name}: {total}日 (契約: {contract}日) {status}")

    # Excel出力
    save_result_to_excel(best_shift, employee_names, holiday, required_workers, output_file)

    print(f"\n適合度推移: {history[0]} → {history[-1]}")
