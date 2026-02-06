"""Excel template generator for shift input data."""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]


@dataclass
class EmployeePreset:
    """Preset employee data for template generation."""

    name: str
    employee_type: str = "正規"  # 正規 / パート
    section: str = ""  # 仕込み / ランチ / 仕込み・ランチ / ホール
    vacation_days: int = 0
    holidays: int = 9
    unavailable_weekdays: list[int] = field(default_factory=list)  # 0=Mon..6=Sun


# Column offset constants for the new layout
# A=社員名, B=雇用形態, C=セクション, D=有休残日数, E onwards=日付
_DAY_COL_OFFSET = 4  # Day 1 starts at column 5 (E)


def generate_template(
    filepath: str | Path,
    year: int,
    month: int,
    num_employees: int = 10,
    default_holidays: int = 9,
    default_required: int = 7,
    employee_names: list[str] | None = None,
    employee_presets: list[EmployeePreset] | None = None,
    kitchen_required: int | None = None,
) -> Path:
    """Generate an Excel template for shift input.

    Args:
        filepath: Output file path.
        year: Target year (e.g. 2026).
        month: Target month (1-12).
        num_employees: Number of employees (default: 10).
        default_holidays: Default holiday count per employee.
        default_required: Default required workers per day.
        employee_names: Optional list of employee names.
        employee_presets: Optional list of EmployeePreset for detailed info.
        kitchen_required: If set, add a kitchen required workers row.

    Returns:
        Path to the generated file.

    The generated layout:
        Row 1: Title
        Row 2: (empty)
        Row 3: Column headers - "社員名", "雇用形態", "セクション", "有休残日数", 1..N, "休日数"
        Row 4: Weekday row  - "曜日", "", "", "", Mon/Tue/...
        Row 5..(5+num_employees-1): Employee rows
        Row (5+num_employees): (empty)
        Row (5+num_employees+1): Required workers (kitchen) - "必要人数（キッチン）", N, N, ...
    """
    filepath = Path(filepath)
    num_days = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = "シフト表"

    # --- Styles ---
    title_font = Font(bold=True, size=14, name="Arial")
    header_font = Font(bold=True, size=11, name="Arial")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill("solid", fgColor="DAEEF3")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    yellow_fill = PatternFill("solid", fgColor="FFFFCC")
    orange_fill = PatternFill("solid", fgColor="FFC000")
    sat_fill = PatternFill("solid", fgColor="CCE5FF")
    sun_fill = PatternFill("solid", fgColor="FFCCCC")
    unavailable_fill = PatternFill("solid", fgColor="D9D9D9")
    weekday_font = Font(size=10, name="Arial")
    sat_font = Font(size=10, name="Arial", color="0000FF")
    sun_font = Font(size=10, name="Arial", color="FF0000")

    # --- Row 1: Title ---
    ws.cell(row=1, column=1, value=f"シフト表（{year}年{month}月）")
    ws.cell(row=1, column=1).font = title_font

    # --- Row 3: Column headers ---
    header_row = 3

    # Fixed columns: A=社員名, B=雇用形態, C=セクション, D=有休残日数
    fixed_headers = [
        ("社員名", blue_fill),
        ("雇用形態", blue_fill),
        ("セクション", blue_fill),
        ("有休残日数", blue_fill),
    ]
    for col_idx, (label, fill) in enumerate(fixed_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

    # Day columns: E onwards
    for d in range(1, num_days + 1):
        col = _DAY_COL_OFFSET + d
        cell = ws.cell(row=header_row, column=col, value=d)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border

        weekday_idx = calendar.weekday(year, month, d)
        if weekday_idx == 5:  # Saturday
            cell.fill = sat_fill
            cell.font = Font(bold=True, size=11, name="Arial", color="0000FF")
        elif weekday_idx == 6:  # Sunday
            cell.fill = sun_fill
            cell.font = Font(bold=True, size=11, name="Arial", color="FF0000")

    holiday_col = _DAY_COL_OFFSET + num_days + 1
    cell = ws.cell(row=header_row, column=holiday_col, value="休日数")
    cell.font = header_font
    cell.fill = yellow_fill
    cell.alignment = center
    cell.border = thin_border

    # --- Row 4: Weekday names ---
    weekday_row = 4

    cell = ws.cell(row=weekday_row, column=1, value="曜日")
    cell.font = weekday_font
    cell.fill = blue_fill
    cell.alignment = center
    cell.border = thin_border

    # Empty cells for B, C, D in weekday row
    for c in range(2, _DAY_COL_OFFSET + 1):
        cell = ws.cell(row=weekday_row, column=c)
        cell.border = thin_border
        cell.fill = blue_fill

    for d in range(1, num_days + 1):
        col = _DAY_COL_OFFSET + d
        weekday_idx = calendar.weekday(year, month, d)
        weekday_name = _WEEKDAY_JA[weekday_idx]

        cell = ws.cell(row=weekday_row, column=col, value=weekday_name)
        cell.alignment = center
        cell.border = thin_border

        if weekday_idx == 5:
            cell.font = sat_font
            cell.fill = sat_fill
        elif weekday_idx == 6:
            cell.font = sun_font
            cell.fill = sun_fill
        else:
            cell.font = weekday_font

    # --- Employee rows ---
    if employee_presets is not None:
        num_employees = len(employee_presets)
    elif employee_names is None:
        employee_names = [f"社員{chr(65 + i)}" for i in range(num_employees)]
    else:
        while len(employee_names) < num_employees:
            employee_names.append(f"社員{len(employee_names) + 1}")
        employee_names = employee_names[:num_employees]

    for emp_idx in range(num_employees):
        row = 5 + emp_idx  # excel 1-indexed

        if employee_presets is not None:
            preset = employee_presets[emp_idx]
            emp_name = preset.name
            emp_type = preset.employee_type
            emp_section = preset.section
            emp_vacation = preset.vacation_days
            emp_holidays = preset.holidays
            unavailable_wdays = preset.unavailable_weekdays
        else:
            emp_name = employee_names[emp_idx]
            emp_type = "正規"
            emp_section = ""
            emp_vacation = 0
            emp_holidays = default_holidays
            unavailable_wdays = []

        # A: Employee name
        cell = ws.cell(row=row, column=1, value=emp_name)
        cell.font = Font(name="Arial", size=11)
        cell.fill = green_fill
        cell.alignment = center
        cell.border = thin_border

        # B: Employment type
        cell = ws.cell(row=row, column=2, value=emp_type)
        cell.font = Font(name="Arial", size=11)
        cell.alignment = center
        cell.border = thin_border

        # C: Section
        cell = ws.cell(row=row, column=3, value=emp_section)
        cell.font = Font(name="Arial", size=11)
        cell.alignment = center
        cell.border = thin_border

        # D: Available vacation days
        cell = ws.cell(row=row, column=4, value=emp_vacation)
        cell.font = Font(name="Arial", size=11)
        cell.alignment = center
        cell.border = thin_border

        # Day cells
        for d in range(1, num_days + 1):
            col = _DAY_COL_OFFSET + d
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.border = thin_border

            weekday_idx = calendar.weekday(year, month, d)

            # Mark unavailable days with ×
            if weekday_idx in unavailable_wdays:
                cell.value = "×"
                cell.fill = unavailable_fill
                cell.font = Font(name="Arial", size=11, bold=True)
            else:
                # Light background for Sat/Sun
                if weekday_idx == 5:
                    cell.fill = PatternFill("solid", fgColor="EBF5FF")
                elif weekday_idx == 6:
                    cell.fill = PatternFill("solid", fgColor="FFF0F0")

        # Holiday count
        cell = ws.cell(row=row, column=holiday_col, value=emp_holidays)
        cell.font = Font(name="Arial", size=11)
        cell.fill = yellow_fill
        cell.alignment = center
        cell.border = thin_border

    # --- Required workers row ---
    req_row = 5 + num_employees + 1  # Skip one empty row

    req_label = "必要人数（キッチン）" if kitchen_required is not None else "必要人数"
    req_count = kitchen_required if kitchen_required is not None else default_required

    cell = ws.cell(row=req_row, column=1, value=req_label)
    cell.font = header_font
    cell.fill = orange_fill
    cell.alignment = center
    cell.border = thin_border

    for d in range(1, num_days + 1):
        col = _DAY_COL_OFFSET + d
        cell = ws.cell(row=req_row, column=col, value=req_count)
        cell.font = Font(name="Arial", size=11)
        cell.fill = orange_fill
        cell.alignment = center
        cell.border = thin_border

    # --- Instructions sheet ---
    ws_help = wb.create_sheet("入力ガイド")
    _write_instructions(ws_help, year, month, num_employees, num_days)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    for col_idx in range(_DAY_COL_OFFSET + 1, holiday_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.5
    ws.column_dimensions[get_column_letter(holiday_col)].width = 8

    # Freeze pane: fix employee info columns and header rows
    freeze_col = get_column_letter(_DAY_COL_OFFSET + 1)
    ws.freeze_panes = f"{freeze_col}5"

    wb.save(str(filepath))
    return filepath


def generate_kimachiya_template(
    filepath: str | Path,
    year: int,
    month: int,
) -> Path:
    """Generate a kimachiya-specific shift template.

    Presets 5 kitchen staff with their employment info and constraints:
    - 川崎聡: 正規, 仕込み
    - 斎藤駿児: 正規, 仕込み・ランチ
    - 平田園美: パート, 仕込み
    - 島村誠: 正規, ランチ (水曜出勤不可)
    - 橋本由紀: パート, ランチ
    """
    presets = [
        EmployeePreset(
            name="川崎聡",
            employee_type="正規",
            section="仕込み",
            vacation_days=10,
            holidays=9,
        ),
        EmployeePreset(
            name="斎藤駿児",
            employee_type="正規",
            section="仕込み・ランチ",
            vacation_days=10,
            holidays=9,
        ),
        EmployeePreset(
            name="平田園美",
            employee_type="パート",
            section="仕込み",
            vacation_days=5,
            holidays=8,
        ),
        EmployeePreset(
            name="島村誠",
            employee_type="正規",
            section="ランチ",
            vacation_days=10,
            holidays=9,
            unavailable_weekdays=[2],  # Wednesday = 水曜日 (index 2)
        ),
        EmployeePreset(
            name="橋本由紀",
            employee_type="パート",
            section="ランチ",
            vacation_days=5,
            holidays=8,
        ),
    ]

    return generate_template(
        filepath=filepath,
        year=year,
        month=month,
        employee_presets=presets,
        kitchen_required=3,
    )


def _write_instructions(ws, year: int, month: int, num_employees: int, num_days: int) -> None:
    """Write the instructions sheet."""
    title_font = Font(bold=True, size=14, name="Arial")
    header_font = Font(bold=True, size=12, name="Arial")
    body_font = Font(size=11, name="Arial")

    instructions = [
        (f"シフト表入力ガイド（{year}年{month}月）", title_font),
        ("", body_font),
        ("【シフト表シートの入力方法】", header_font),
        ("", body_font),
        ("1. 社員情報（A〜D列）", header_font),
        ("   A列: 社員名", body_font),
        ("   B列: 雇用形態（正規 / パート）", body_font),
        ("   C列: セクション（仕込み / ランチ / 仕込み・ランチ / ホール）", body_font),
        ("   D列: 有給休暇取得可能日数", body_font),
        ("", body_font),
        ("2. 希望休の入力（E列〜）", header_font),
        ("   希望休のセルに「◎」を入力してください。", body_font),
        ("   ◎のセルはGAが変更しない固定休日（原則有給）として扱われます。", body_font),
        ("   空欄のセルはGAがスケジューリング対象とします。", body_font),
        ("   「×」は出勤不可日を示します（GAは変更しません）。", body_font),
        ("", body_font),
        ("3. 休日数", header_font),
        ("   最終列に各社員の契約休日数を入力してください。", body_font),
        ("   希望休（◎）と出勤不可（×）の数もこの休日数に含まれます。", body_font),
        ("   例: 休日数=9、◎が2個、×が4個 → GAが残り3日の休日を自動配置", body_font),
        ("", body_font),
        ("4. 必要人数", header_font),
        ("   最下行に各日の必要出勤人数を入力してください。", body_font),
        ("   日によって必要人数が異なる場合は個別に変更してください。", body_font),
        ("", body_font),
        ("【セルの値】", header_font),
        ("   空欄 = 出勤可能（GAがスケジューリング）", body_font),
        ("   ◎   = 希望休（固定、原則有給。GAは変更しない）", body_font),
        ("   ×   = 出勤不可（完全固定。通院日・非勤務曜日等）", body_font),
        ("   ※ GAが休日を割り当てると「休」になります", body_font),
        ("", body_font),
        ("【注意事項】", header_font),
        ("   - シート名は「シフト表」のまま変更しないでください", body_font),
        ("   - 行・列の挿入・削除はしないでください", body_font),
        (f"   - 日数は{num_days}日（{year}年{month}月）です", body_font),
        ("   - 休日数は希望休・出勤不可を含む合計数です", body_font),
    ]

    for row_idx, (text, font) in enumerate(instructions, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = font

    ws.column_dimensions["A"].width = 80


def main() -> None:
    """CLI entry point for template generation."""
    import argparse

    parser = argparse.ArgumentParser(description="シフト表テンプレート生成")
    parser.add_argument("--year", type=int, required=True, help="年（例: 2026）")
    parser.add_argument("--month", type=int, required=True, help="月（1-12）")
    parser.add_argument("--employees", type=int, default=5, help="社員数（デフォルト: 5）")
    parser.add_argument("--holidays", type=int, default=9, help="デフォルト休日数")
    parser.add_argument("--required", type=int, default=3, help="デフォルト必要出勤人数")
    parser.add_argument(
        "--kimachiya", action="store_true", help="木町家プリセットで生成"
    )
    parser.add_argument("-o", "--output", default=None, help="出力ファイルパス")
    args = parser.parse_args()

    if args.output is None:
        args.output = f"shift_input_{args.year}_{args.month:02d}.xlsx"

    if args.kimachiya:
        filepath = generate_kimachiya_template(
            filepath=args.output,
            year=args.year,
            month=args.month,
        )
    else:
        filepath = generate_template(
            filepath=args.output,
            year=args.year,
            month=args.month,
            num_employees=args.employees,
            default_holidays=args.holidays,
            default_required=args.required,
        )
    print(f"テンプレート生成完了: {filepath}")


if __name__ == "__main__":
    main()
