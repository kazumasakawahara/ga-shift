"""Excel writer for shift results."""

from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ga_shift.models.employee import Section
from ga_shift.models.schedule import ShiftInput, ShiftResult
from ga_shift.models.validation import ValidationReport

# Column layout: A=社員名, B=雇用形態, C=セクション, D=有休残, E onwards=days
_DAY_COL_OFFSET = 4  # 0-indexed: day 1 starts at column index 4 (openpyxl col 5)


def write_result_excel(
    filepath: str | Path,
    shift_result: ShiftResult,
    shift_input: ShiftInput,
    validation_report: ValidationReport | None = None,
) -> None:
    """Write GA result to Excel file."""
    filepath = Path(filepath)
    wb = Workbook()

    _write_schedule_sheet(wb.active, shift_result, shift_input)
    if validation_report:
        _write_validation_sheet(wb, validation_report)

    wb.save(str(filepath))


def _write_schedule_sheet(
    ws, shift_result: ShiftResult, shift_input: ShiftInput
) -> None:
    ws.title = "GA結果シフト表"

    header_font = Font(bold=True, size=11, name="Arial")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill("solid", fgColor="DAEEF3")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    red_font = Font(color="FF0000", bold=True, name="Arial")
    gray_font = Font(color="808080", bold=True, name="Arial")
    work_fill = PatternFill("solid", fgColor="FFFFFF")
    holiday_fill = PatternFill("solid", fgColor="D9E2F3")
    preferred_fill = PatternFill("solid", fgColor="FCE4EC")
    unavailable_fill = PatternFill("solid", fgColor="D9D9D9")

    schedule = shift_result.best_schedule
    num_employees = shift_input.num_employees
    num_days = shift_input.num_days

    # Title
    ws.cell(row=1, column=1, value="GA最適化シフト表")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, name="Arial")

    # Header row
    fixed_headers = [
        ("社員名", blue_fill),
        ("雇用形態", blue_fill),
        ("セクション", blue_fill),
        ("有休残", blue_fill),
    ]
    for col_idx, (label, fill) in enumerate(fixed_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = border

    for d in range(num_days):
        cell = ws.cell(row=3, column=_DAY_COL_OFFSET + d + 1, value=d + 1)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    # Summary columns after day data
    col_actual = _DAY_COL_OFFSET + num_days + 1
    col_contract = col_actual + 1
    col_vacation_used = col_contract + 1
    for col, label in [
        (col_actual, "実休日"),
        (col_contract, "契約"),
        (col_vacation_used, "有給消化"),
    ]:
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = center
        cell.border = border

    # Data rows
    label_map = {0: "出", 1: "休", 2: "◎", 3: "×"}

    for emp in shift_input.employees:
        row_num = 4 + emp.index

        # A: Name
        ws.cell(row=row_num, column=1, value=emp.name)
        ws.cell(row=row_num, column=1).font = Font(name="Arial", size=11)
        ws.cell(row=row_num, column=1).fill = green_fill
        ws.cell(row=row_num, column=1).alignment = center
        ws.cell(row=row_num, column=1).border = border

        # B: Employee type
        emp_type_str = emp.employee_type.value if emp.employee_type else ""
        ws.cell(row=row_num, column=2, value=emp_type_str)
        ws.cell(row=row_num, column=2).alignment = center
        ws.cell(row=row_num, column=2).border = border

        # C: Section
        section_str = emp.section.value if emp.section else ""
        ws.cell(row=row_num, column=3, value=section_str)
        ws.cell(row=row_num, column=3).alignment = center
        ws.cell(row=row_num, column=3).border = border

        # D: Available vacation days
        ws.cell(row=row_num, column=4, value=emp.available_vacation_days)
        ws.cell(row=row_num, column=4).alignment = center
        ws.cell(row=row_num, column=4).border = border

        actual_holidays = 0
        vacation_used = 0  # Count of ◎ (preferred off / paid leave)
        for d in range(num_days):
            val = int(schedule[emp.index, d])
            cell = ws.cell(row=row_num, column=_DAY_COL_OFFSET + d + 1)
            cell.value = label_map.get(val, str(val))
            cell.alignment = center
            cell.border = border

            if val == 0:
                cell.fill = work_fill
            elif val == 1:
                cell.fill = holiday_fill
                actual_holidays += 1
            elif val == 2:
                cell.fill = preferred_fill
                cell.font = red_font
                actual_holidays += 1
                vacation_used += 1
            elif val == 3:
                cell.fill = unavailable_fill
                cell.font = gray_font
                actual_holidays += 1

        # Summary columns
        ws.cell(row=row_num, column=col_actual, value=actual_holidays)
        ws.cell(row=row_num, column=col_actual).alignment = center
        ws.cell(row=row_num, column=col_actual).border = border

        cell = ws.cell(row=row_num, column=col_contract, value=emp.required_holidays)
        cell.alignment = center
        cell.border = border
        cell.fill = yellow_fill

        # Vacation usage: mark red if over limit
        cell = ws.cell(row=row_num, column=col_vacation_used, value=vacation_used)
        cell.alignment = center
        cell.border = border
        if vacation_used > emp.available_vacation_days:
            cell.fill = PatternFill("solid", fgColor="FF9999")
            cell.font = Font(color="FF0000", bold=True, name="Arial")

    # --- Kitchen worker count summary ---
    # Determine which employees are "kitchen" (PREP, LUNCH, or PREP_LUNCH)
    kitchen_sections = {Section.PREP, Section.LUNCH, Section.PREP_LUNCH}
    kitchen_indices = [
        emp.index
        for emp in shift_input.employees
        if emp.section in kitchen_sections
    ]

    summary_row = 4 + num_employees + 1

    # Kitchen workers row
    ws.cell(row=summary_row, column=1, value="キッチン出勤")
    ws.cell(row=summary_row, column=1).font = header_font
    ws.cell(row=summary_row, column=1).fill = PatternFill("solid", fgColor="FFC000")
    ws.cell(row=summary_row, column=1).alignment = center
    ws.cell(row=summary_row, column=1).border = border

    binary = np.where(schedule >= 2, 1, schedule)
    for d in range(num_days):
        if kitchen_indices:
            kitchen_workers = sum(
                1 for idx in kitchen_indices if binary[idx, d] == 0
            )
        else:
            # Fallback: count all workers
            holiday_count = int(np.sum(binary[:, d]))
            kitchen_workers = num_employees - holiday_count

        cell = ws.cell(row=summary_row, column=_DAY_COL_OFFSET + d + 1, value=kitchen_workers)
        cell.alignment = center
        cell.border = border

        # Color based on target (use kitchen requirement or overall)
        target = 3  # Default kitchen target
        if shift_input.required_kitchen_workers is not None:
            target = int(shift_input.required_kitchen_workers[d])
        elif shift_input.required_workers is not None:
            target = int(shift_input.required_workers[d])

        if kitchen_workers < target:
            cell.fill = PatternFill("solid", fgColor="FF9999")
        elif kitchen_workers > target:
            cell.fill = PatternFill("solid", fgColor="FFFF99")
        else:
            cell.fill = PatternFill("solid", fgColor="99FF99")

    # Required workers row
    req_row = summary_row + 1
    req_label = "必要人数（キッチン）" if shift_input.required_kitchen_workers is not None else "必要人数"
    ws.cell(row=req_row, column=1, value=req_label)
    ws.cell(row=req_row, column=1).font = header_font
    ws.cell(row=req_row, column=1).fill = PatternFill("solid", fgColor="FFC000")
    ws.cell(row=req_row, column=1).alignment = center
    ws.cell(row=req_row, column=1).border = border

    req_source = shift_input.required_kitchen_workers if shift_input.required_kitchen_workers is not None else shift_input.required_workers
    for d in range(num_days):
        cell = ws.cell(row=req_row, column=_DAY_COL_OFFSET + d + 1, value=int(req_source[d]))
        cell.alignment = center
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFC000")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    for col_idx in range(_DAY_COL_OFFSET + 1, col_vacation_used + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    ws.column_dimensions[get_column_letter(col_actual)].width = 8
    ws.column_dimensions[get_column_letter(col_contract)].width = 6
    ws.column_dimensions[get_column_letter(col_vacation_used)].width = 8


def _write_validation_sheet(wb: Workbook, report: ValidationReport) -> None:
    ws = wb.create_sheet("バリデーション結果")

    header_font = Font(bold=True, size=11, name="Arial")

    ws.cell(row=1, column=1, value="バリデーション結果")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, name="Arial")

    ws.cell(row=3, column=1, value=f"合計ペナルティ: {report.total_penalty:.1f}")
    ws.cell(row=4, column=1, value=f"エラー数: {report.error_count}")
    ws.cell(row=5, column=1, value=f"警告数: {report.warning_count}")

    # Constraint scores table
    row = 7
    for col, header in enumerate(["制約ID", "ペナルティ", "違反数"], 1):
        ws.cell(row=row, column=col, value=header).font = header_font

    for cs in report.constraint_scores:
        row += 1
        ws.cell(row=row, column=1, value=cs.constraint_id)
        ws.cell(row=row, column=2, value=cs.penalty)
        ws.cell(row=row, column=3, value=len(cs.violations))

    # Violations detail
    row += 2
    ws.cell(row=row, column=1, value="違反詳細").font = header_font
    row += 1
    for col, header in enumerate(["制約ID", "重要度", "メッセージ"], 1):
        ws.cell(row=row, column=col, value=header).font = header_font

    for v in report.violations:
        row += 1
        ws.cell(row=row, column=1, value=v.constraint_id)
        ws.cell(row=row, column=2, value=v.severity.value)
        ws.cell(row=row, column=3, value=v.message)
